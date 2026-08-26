"""Generación del archivo Excel para casos de prueba AER."""

from __future__ import annotations

from collections.abc import Sequence
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.worksheet.datavalidation import DataValidation

from apps.aer_test_case.schemas.test_case_response import (
    AerTestCaseResponse,
)
from apps.aer_test_case.services.excel_row_builder import (
    AerExcelRow,
)
from apps.aer_test_case.services.excel_row_builder import (
    build_excel_rows,
)


SHEET_NAME = "UAT Scenarios"

HEADERS = (
    "Scenario ID",
    "Description",
    "FDD reference",
    "Expected Result",
    "Exception text if applicable",
    "Input",
    "Comments",
    "Test Status",
    "Date when tested",
    "Prioridad",
    "Tester",
    "Bug asociado",
)

COLUMN_WIDTHS = (
    14,
    60,
    20,
    55,
    45,
    30,
    35,
    18,
    20,
    12,
    22,
    20,
)

HEADER_FILL = "205327"
HEADER_FONT_COLOR = "FFFFFF"
BORDER_COLOR = "B7B7B7"

TEST_STATUS_OPTIONS = (
    "Not Tested",
    "Success",
    "Fail",
)

TEST_STATUS_COLUMN = "H"

def generate_aer_xlsx(
    responses: Sequence[AerTestCaseResponse],
) -> bytes:
    """Genera el XLSX final de casos de prueba AER."""
    rows = build_excel_rows(
        responses
    )

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = SHEET_NAME

    _write_headers(
        worksheet
    )

    _write_rows(
        worksheet=worksheet,
        rows=rows,
    )

    _apply_test_status_validation(
        worksheet=worksheet,
        row_count=len(rows),
    )

    _apply_sheet_format(
        worksheet
    )

    output = BytesIO()

    workbook.save(
        output
    )

    return output.getvalue()


def _write_headers(
    worksheet,
) -> None:
    """Escribe las columnas oficiales del Excel AER."""
    for column_index, header in enumerate(
        HEADERS,
        start=1,
    ):
        worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )


def _write_rows(
    *,
    worksheet,
    rows: Sequence[AerExcelRow],
) -> None:
    """Escribe los casos generados en la hoja AER."""
    for row_index, row in enumerate(
        rows,
        start=2,
    ):
        values = (
            row.scenario_id,
            row.description,
            row.fdd_reference,
            row.expected_result,
            row.exception_text,
            row.input_value,
            row.comments,
            row.test_status,
            row.date_when_tested,
            row.priority,
            row.tester,
            row.associated_bug,
        )

        for column_index, value in enumerate(
            values,
            start=1,
        ):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

def _apply_test_status_validation(
    *,
    worksheet,
    row_count: int,
) -> None:
    """Agrega las opciones permitidas para Test Status."""
    if row_count <= 0:
        return

    options = ",".join(
        TEST_STATUS_OPTIONS
    )

    validation = DataValidation(
        type="list",
        formula1=f'"{options}"',
        allow_blank=False,
    )

    validation.error = (
        "Selecciona un Test Status válido."
    )

    validation.errorTitle = (
        "Test Status inválido"
    )

    validation.prompt = (
        "Selecciona Not Tested, Success o Fail."
    )

    validation.promptTitle = "Test Status"

    validation.showErrorMessage = True
    validation.showInputMessage = True

    worksheet.add_data_validation(
        validation
    )

    first_row = 2

    last_row = (
        row_count + 1
    )

    validation.add(
        f"{TEST_STATUS_COLUMN}"
        f"{first_row}:"
        f"{TEST_STATUS_COLUMN}"
        f"{last_row}"
    )

def _apply_sheet_format(
    worksheet,
) -> None:
    """Aplica formato básico para facilitar la lectura."""
    thin_border = Side(
        style="thin",
        color=BORDER_COLOR,
    )

    border = Border(
        left=thin_border,
        right=thin_border,
        top=thin_border,
        bottom=thin_border,
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor=HEADER_FILL,
    )

    header_font = Font(
        bold=True,
        color=HEADER_FONT_COLOR,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = border

    for column_index, width in enumerate(
        COLUMN_WIDTHS,
        start=1,
    ):
        column_letter = worksheet.cell(
            row=1,
            column=column_index,
        ).column_letter

        worksheet.column_dimensions[
            column_letter
        ].width = width

    for row in worksheet.iter_rows(
        min_row=2,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
            cell.border = border

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )
    


    worksheet.row_dimensions[1].height = 35