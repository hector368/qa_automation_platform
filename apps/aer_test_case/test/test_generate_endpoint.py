"""Prueba HTTP real de generación y descarga AER."""

from __future__ import annotations

import os
from io import BytesIO
from pathlib import Path

import django


os.environ.setdefault(
    "DJANGO_SETTINGS_MODULE",
    "config.settings.local",
)

django.setup()


from django.core.files.uploadedfile import (
    SimpleUploadedFile,
)
from django.test import Client
from django.urls import reverse
from openpyxl import load_workbook


TEST_DIRECTORY = Path(__file__).resolve().parent

TEST_FILE_PATH = (
    TEST_DIRECTORY
    / "files"
    / "mcc_015_fdd.pdf"
)

SELECTED_REQUIREMENT = "MCC.015.001"


def run_test() -> None:
    """Genera un RQ por HTTP y descarga su XLSX."""
    if not TEST_FILE_PATH.is_file():
        raise FileNotFoundError(
            f"Test document was not found: "
            f"{TEST_FILE_PATH}"
        )

    uploaded_file = SimpleUploadedFile(
        name=TEST_FILE_PATH.name,
        content=TEST_FILE_PATH.read_bytes(),
        content_type="application/pdf",
    )

    client = Client(
        HTTP_HOST="localhost",
    )

    print()
    print(
        f"Generando: {SELECTED_REQUIREMENT}"
    )
    print()

    response = client.post(
        reverse(
            "aer_test_case:generate"
        ),
        {
            "document": uploaded_file,
            "selected_requirements": (
                SELECTED_REQUIREMENT
            ),
        },
    )

    assert response.status_code == 200

    payload = response.json()

    assert payload["ok"] is True

    assert (
        payload["selected_requirements"]
        == [
            SELECTED_REQUIREMENT,
        ]
    )

    assert payload["total_test_cases"] > 0
    assert payload["result_id"]
    assert payload["download_url"]

    print(
        f"Casos generados: "
        f"{payload['total_test_cases']}"
    )

    print(
        f"Archivo: {payload['filename']}"
    )

    print(
        f"Download URL: "
        f"{payload['download_url']}"
    )

    print(
        f"Tokens: {payload['usage']}"
    )

    print(
        f"Costo: {payload['cost']}"
    )

    download_response = client.get(
        payload["download_url"],
    )

    assert download_response.status_code == 200

    assert (
        download_response[
            "Content-Type"
        ]
        == (
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    workbook = load_workbook(
        BytesIO(
            download_response.content
        )
    )

    assert (
        "UAT Scenarios"
        in workbook.sheetnames
    )

    worksheet = workbook[
        "UAT Scenarios"
    ]

    assert worksheet.max_row > 1
    assert worksheet.max_column == 12

    reference = worksheet.cell(
        row=2,
        column=3,
    ).value

    assert reference == (
        "Referencia al requerimiento: "
        "MCC.015.001 - "
        "Download the MAE catalog"
    )

    print()
    print(
        f"Filas descargadas: "
        f"{worksheet.max_row - 1}"
    )

    print()
    print(
        "AER generate endpoint test: OK"
    )


if __name__ == "__main__":
    run_test()