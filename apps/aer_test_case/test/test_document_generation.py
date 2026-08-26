"""Prueba local del resultado final del orquestador AER."""

from io import BytesIO

from openpyxl import load_workbook

from apps.aer_test_case.schemas.test_case_response import (
    AerTestCaseResponse,
)
from apps.aer_test_case.services.orchestrator import (
    build_document_generation,
)
from apps.test_cases.services.token_usage import (
    TokenUsage,
)


def build_sample_response() -> AerTestCaseResponse:
    """Construye una respuesta AER válida para la prueba."""
    data = {
        "requirement_id": "MCC.015.001",
        "requirement_title": "Download the MAE catalog",
        "is_testable": True,
        "not_testable_reason": None,
        "test_cases": [
            {
                "description": (
                    "Precondiciones:\n"
                    "No especificadas.\n\n"
                    "Tipo de flujo:\n"
                    "Happy Path\n\n"
                    "Descripción:\n"
                    "Validar la descarga del catálogo.\n\n"
                    "Pasos:\n"
                    "1. Iniciar el proceso.\n"
                    "2. Descargar el catálogo."
                ),
                "expected_result": (
                    "El catálogo es descargado."
                ),
                "exception_text": None,
                "input": "Catálogo MAE Routes",
                "comments": None,
                "priority": 1,
            },
            {
                "description": (
                    "Precondiciones:\n"
                    "No especificadas.\n\n"
                    "Tipo de flujo:\n"
                    "Excepción\n\n"
                    "Descripción:\n"
                    "Validar catálogo no disponible.\n\n"
                    "Pasos:\n"
                    "1. Iniciar el proceso.\n"
                    "2. Intentar localizar el catálogo."
                ),
                "expected_result": (
                    "El BOT suspende la ejecución."
                ),
                "exception_text": (
                    "El catálogo no está disponible."
                ),
                "input": "Catálogo MAE Routes",
                "comments": None,
                "priority": 2,
            },
        ],
    }

    return AerTestCaseResponse.model_validate(
        data
    )


def run_test() -> None:
    """Valida respuestas, métricas y XLSX final."""
    response = build_sample_response()

    usage = TokenUsage(
        input_tokens=1000,
        output_tokens=500,
    )

    generation = build_document_generation(
        responses=[
            response,
        ],
        usage=usage,
    )

    assert len(generation.responses) == 1

    assert generation.total_test_cases == 2

    assert generation.usage.input_tokens == 1000

    assert generation.usage.output_tokens == 500

    assert generation.usage.total_tokens == 1500

    assert generation.xlsx_bytes

    workbook = load_workbook(
        BytesIO(
            generation.xlsx_bytes
        )
    )

    worksheet = workbook[
        "UAT Scenarios"
    ]

    assert worksheet.max_row == 3
    assert worksheet.max_column == 12

    assert (
        worksheet.cell(
            row=2,
            column=1,
        ).value
        == 1
    )

    assert (
        worksheet.cell(
            row=3,
            column=1,
        ).value
        == 2
    )

    first_reference = worksheet.cell(
        row=2,
        column=3,
    ).value

    assert first_reference == (
        "Referencia al requerimiento: "
        "MCC.015.001 - "
        "Download the MAE catalog"
    )

    assert (
        worksheet.cell(
            row=2,
            column=10,
        ).value
        == 1
    )

    assert (
        worksheet.cell(
            row=3,
            column=10,
        ).value
        == 2
    )

    print(
        "AER document generation test: OK"
    )


if __name__ == "__main__":
    run_test()