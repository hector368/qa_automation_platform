"""Prueba local del generador XLSX AER."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from apps.aer_test_case.schemas.test_case_response import (
    AerTestCaseResponse,
)
from apps.aer_test_case.services.xlsx_generator import (
    HEADERS,
)
from apps.aer_test_case.services.xlsx_generator import (
    SHEET_NAME,
)
from apps.aer_test_case.services.xlsx_generator import (
    generate_aer_xlsx,
)


TEST_DIRECTORY = Path(__file__).resolve().parent

OUTPUT_FILE_PATH = (
    TEST_DIRECTORY
    / "aer_test_cases_result.xlsx"
)


def build_test_responses() -> tuple[
    AerTestCaseResponse,
    ...,
]:
    """Construye respuestas controladas para probar el XLSX."""
    response = AerTestCaseResponse.model_validate(
        {
            "requirement_id": "MCC.015.001",
            "requirement_title": (
                "Download the MAE catalog"
            ),
            "is_testable": True,
            "not_testable_reason": None,
            "test_cases": [
                {
                    "description": (
                        "Precondiciones:\n"
                        "No especificadas.\n\n"
                        "Tipo de flujo:\n"
                        "Happy Path\n\n"
                        "Descripción:\n"
                        "Validar la descarga del catálogo.\n\n"
                        "Pasos:\n"
                        "1. Iniciar el proceso.\n"
                        "2. Descargar el catálogo."
                    ),
                    "expected_result": (
                        "El catálogo MAE Routes "
                        "es descargado."
                    ),
                    "exception_text": None,
                    "input": "Catálogo MAE Routes",
                    "comments": None,
                    "priority": 1,
                },
                {
                    "description": (
                        "Precondiciones:\n"
                        "No especificadas.\n\n"
                        "Tipo de flujo:\n"
                        "Excepción\n\n"
                        "Descripción:\n"
                        "Validar el comportamiento cuando "
                        "el catálogo no está disponible.\n\n"
                        "Pasos:\n"
                        "1. Iniciar el proceso.\n"
                        "2. Intentar localizar el catálogo."
                    ),
                    "expected_result": (
                        "El BOT suspende la ejecución."
                    ),
                    "exception_text": (
                        "El catálogo MAE Routes "
                        "no está disponible."
                    ),
                    "input": "Catálogo MAE Routes",
                    "comments": None,
                    "priority": 2,
                },
            ],
        }
    )

    return (
        response,
    )


def run_test() -> None:
    """Genera y valida un Excel AER controlado."""
    responses = build_test_responses()

    expected_test_cases = sum(
        len(response.test_cases)
        for response in responses
    )

    print()
    print(
        f"Respuestas cargadas: "
        f"{len(responses)}"
    )

    print(
        f"Casos esperados: "
        f"{expected_test_cases}"
    )

    xlsx_bytes = generate_aer_xlsx(
        responses
    )

    assert xlsx_bytes

    OUTPUT_FILE_PATH.write_bytes(
        xlsx_bytes
    )

    workbook = load_workbook(
        BytesIO(xlsx_bytes)
    )

    assert SHEET_NAME in workbook.sheetnames

    worksheet = workbook[
        SHEET_NAME
    ]

    header_color = (
        worksheet["A1"]
        .fill
        .fgColor
        .rgb
    )

    assert header_color is not None

    assert header_color.endswith(
        "205327"
    )

    assert (
        worksheet["H2"].value
        == "Not Tested"
    )

    assert (
        worksheet["H3"].value
        == "Not Tested"
    )

    validations = list(
        worksheet
        .data_validations
        .dataValidation
    )

    assert len(validations) == 1

    status_validation = validations[0]

    assert (
        status_validation.formula1
        == '"Not Tested,Success,Fail"'
    )

    generated_headers = tuple(
        worksheet.cell(
            row=1,
            column=column_index,
        ).value
        for column_index in range(
            1,
            len(HEADERS) + 1,
        )
    )

    assert generated_headers == HEADERS

    assert (
        worksheet.max_row
        == expected_test_cases + 1
    )

    assert worksheet.max_column == 12

    for row_index in range(
        2,
        worksheet.max_row + 1,
    ):
        expected_scenario_id = (
            row_index - 1
        )

        assert (
            worksheet.cell(
                row=row_index,
                column=1,
            ).value
            == expected_scenario_id
        )

        priority = worksheet.cell(
            row=row_index,
            column=10,
        ).value

        assert priority in (
            1,
            2,
        )

        fdd_reference = worksheet.cell(
            row=row_index,
            column=3,
        ).value

        assert fdd_reference == (
            "Referencia al requerimiento: "
            "MCC.015.001 - "
            "Download the MAE catalog"
        )

    print()
    print(
        f"Filas generadas: "
        f"{worksheet.max_row - 1}"
    )

    print(
        f"Columnas: "
        f"{worksheet.max_column}"
    )

    print()
    print(
        "Color encabezado: #205327"
    )

    print(
        "Test Status: "
        "Not Tested / Success / Fail"
    )

    print()
    print(
        "Archivo generado:"
    )

    print(
        OUTPUT_FILE_PATH
    )

    print()
    print(
        "AER XLSX generator test: OK"
    )


if __name__ == "__main__":
    run_test()