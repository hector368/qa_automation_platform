"""Prueba HTTP del streaming AER sin consumir Claude."""

from __future__ import annotations

import json
import os
from io import BytesIO
from unittest.mock import patch

import django


os.environ.setdefault(
    "DJANGO_SETTINGS_MODULE",
    "config.settings.local",
)

django.setup()


from django.core.files.uploadedfile import (
    SimpleUploadedFile,
)
from django.test import Client
from django.urls import reverse
from openpyxl import load_workbook

from apps.aer_test_case.schemas.requirement_schema import (
    RequirementSegment,
)
from apps.aer_test_case.schemas.test_case_response import (
    AerTestCaseResponse,
)
from apps.aer_test_case.services.orchestrator import (
    PreparedAerDocument,
)
from apps.aer_test_case.services.orchestrator import (
    build_document_generation,
)
from apps.test_cases.services.token_usage import (
    TokenUsage,
)


def build_prepared_document() -> PreparedAerDocument:
    """Construye un documento preparado controlado."""
    return PreparedAerDocument(
        filename="sample.pdf",
        requirements=(
            RequirementSegment(
                requirement_id="MCC.015.001",
                title="Download the MAE catalog",
                content=(
                    "MCC.015.001 "
                    "Download the MAE catalog"
                ),
            ),
        ),
    )


def build_generation():
    """Construye un resultado final sin llamar a Claude."""
    response = AerTestCaseResponse.model_validate(
        {
            "requirement_id": "MCC.015.001",
            "requirement_title": (
                "Download the MAE catalog"
            ),
            "is_testable": True,
            "not_testable_reason": None,
            "test_cases": [
                {
                    "description": (
                        "Precondiciones:\n"
                        "No especificadas.\n\n"
                        "Tipo de flujo:\n"
                        "Happy Path\n\n"
                        "Descripción:\n"
                        "Validar la descarga.\n\n"
                        "Pasos:\n"
                        "1. Iniciar."
                    ),
                    "expected_result": (
                        "El catálogo es descargado."
                    ),
                    "exception_text": None,
                    "input": "Catálogo MAE Routes",
                    "comments": None,
                    "priority": 1,
                },
            ],
        }
    )

    return build_document_generation(
        responses=[
            response,
        ],
        usage=TokenUsage(
            input_tokens=100,
            output_tokens=50,
        ),
    )


def fake_iterator(
    *,
    prepared_document,
    selected_requirement_ids,
):
    """Simula eventos del orquestador."""
    del prepared_document

    yield {
        "type": "started",
        "ok": True,
        "total_requirements": 1,
        "selected_requirements": (
            selected_requirement_ids
        ),
        "progress": 0,
    }

    yield {
        "type": "requirement_started",
        "ok": True,
        "requirement_id": "MCC.015.001",
        "requirement_title": (
            "Download the MAE catalog"
        ),
        "current": 1,
        "total": 1,
        "progress": 0,
    }

    yield {
        "type": "requirement_completed",
        "ok": True,
        "requirement_id": "MCC.015.001",
        "requirement_title": (
            "Download the MAE catalog"
        ),
        "generated_test_cases": 1,
        "current": 1,
        "total": 1,
        "progress": 100,
    }

    yield {
        "type": "completed",
        "ok": True,
        "progress": 100,
        "generation": build_generation(),
    }


def run_test() -> None:
    """Valida progreso, caché y descarga XLSX."""
    client = Client(
        HTTP_HOST="localhost",
    )

    uploaded_file = SimpleUploadedFile(
        name="sample.pdf",
        content=b"fake-pdf",
        content_type="application/pdf",
    )

    with patch(
        "apps.aer_test_case.views.prepare_document",
        return_value=build_prepared_document(),
    ):
        with patch(
            "apps.aer_test_case.views."
            "iter_generate_document",
            side_effect=fake_iterator,
        ):
            response = client.post(
                reverse(
                    "aer_test_case:generate_stream"
                ),
                {
                    "document": uploaded_file,
                    "selected_requirements": (
                        "MCC.015.001"
                    ),
                },
            )

            assert response.status_code == 200
            assert response.streaming

            raw_content = b"".join(
                response.streaming_content
            ).decode(
                "utf-8"
            )

    events = [
        json.loads(line)
        for line in raw_content.splitlines()
        if line.strip()
    ]

    assert len(events) == 4

    assert events[0]["type"] == "started"

    assert (
        events[1]["type"]
        == "requirement_started"
    )

    assert (
        events[2]["type"]
        == "requirement_completed"
    )

    final_event = events[3]

    assert (
        final_event["type"]
        == "completed"
    )

    assert final_event["ok"] is True

    assert final_event["progress"] == 100

    assert (
        final_event["total_test_cases"]
        == 1
    )

    assert final_event["download_url"]

    download_response = client.get(
        final_event["download_url"]
    )

    assert (
        download_response.status_code
        == 200
    )

    workbook = load_workbook(
        BytesIO(
            download_response.content
        )
    )

    worksheet = workbook[
        "UAT Scenarios"
    ]

    assert worksheet.max_row == 2
    assert worksheet.max_column == 12

    print()
    print(
        "Eventos recibidos:"
    )

    for event in events:
        print(
            event["type"],
            "|",
            event.get(
                "progress"
            ),
        )

    print()
    print(
        "AER stream endpoint test: OK"
    )


if __name__ == "__main__":
    run_test()