from io import BytesIO

from django.test import SimpleTestCase
from openpyxl import load_workbook

from apps.test_cases.services.ado_rows_builder import (
    ADO_COLUMNS,
    ADO_NCOLS,
)
from apps.test_cases.services.xlsx_generator import (
    WORKSHEET_NAME,
    generate_xlsx,
)


class XlsxGeneratorTests(SimpleTestCase):
    """Pruebas para la generación del archivo XLSX."""

    def test_generates_valid_xlsx(
        self,
    ) -> None:
        """Genera un libro de Excel válido."""
        rows = [
            self._build_metadata_row(),
        ]

        content = generate_xlsx(
            rows
        )

        self.assertTrue(
            content.startswith(
                b"PK"
            )
        )

        self.assertGreater(
            len(content),
            0,
        )

    def test_writes_ado_header(
        self,
    ) -> None:
        """Escribe las quince columnas ADO."""
        content = generate_xlsx(
            [
                self._build_metadata_row(),
            ]
        )

        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
        )

        worksheet = workbook[
            WORKSHEET_NAME
        ]

        header = [
            cell.value
            for cell in worksheet[1]
        ]

        self.assertEqual(
            header,
            list(ADO_COLUMNS),
        )

        workbook.close()

    def test_writes_test_case_data(
        self,
    ) -> None:
        """Escribe metadata y pasos sin modificar contenido."""
        metadata = self._build_metadata_row()

        step = [""] * ADO_NCOLS
        step[3] = "1"
        step[4] = (
            "Validar que el bot procese el archivo"
        )
        step[5] = (
            "El archivo es procesado"
        )

        content = generate_xlsx(
            [
                metadata,
                step,
            ]
        )

        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
        )

        worksheet = workbook[
            WORKSHEET_NAME
        ]

        self.assertEqual(
            worksheet["C2"].value,
            "CFC.003.001.001",
        )

        self.assertEqual(
            worksheet["D3"].value,
            "1",
        )

        self.assertEqual(
            worksheet["E3"].value,
            (
                "Validar que el bot procese "
                "el archivo"
            ),
        )

        workbook.close()

    def test_uses_single_worksheet(
        self,
    ) -> None:
        """Genera únicamente la hoja de casos de prueba."""
        content = generate_xlsx(
            [
                self._build_metadata_row(),
            ]
        )

        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
        )

        self.assertEqual(
            workbook.sheetnames,
            [
                WORKSHEET_NAME,
            ],
        )

        workbook.close()

    def test_rejects_invalid_row_length(
        self,
    ) -> None:
        """Rechaza filas que no contienen quince columnas."""
        invalid_row = [
            ""
        ] * (
            ADO_NCOLS - 1
        )

        with self.assertRaises(
            ValueError,
        ):
            generate_xlsx(
                [
                    invalid_row,
                ]
            )

    @staticmethod
    def _build_metadata_row() -> list[str]:
        """Construye una fila ADO válida para pruebas."""
        row = [""] * ADO_NCOLS

        row[1] = "Test Case"
        row[2] = "CFC.003.001.001"
        row[6] = "Functional"
        row[7] = "1"
        row[8] = "El archivo es procesado."
        row[9] = (
            "Que el bot procese el archivo."
        )
        row[10] = (
            "(Happy Path) - Procesar archivo"
        )
        row[12] = "Design"
        row[13] = "CFC.003"
        row[14] = "Usuario QA"

        return row